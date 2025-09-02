import streamlit as st
from io import BytesIO
import pandas as pd
import time
import altair as alt
import re, io
from typing import Tuple
import collections.abc as abc
import numpy as np
import unicodedata, re
import streamlit_authenticator as stauth
st.set_page_config(page_title="売上ダッシュボード", layout="wide")
def to_dict(obj):
    """Mapping を再帰的に普通の dict へ"""
    if isinstance(obj, abc.Mapping):
        return {k: to_dict(v) for k, v in obj.items()}
    return obj

# ─── 認証 ───
# 認証ブロックを書き換え
if "auth_ok" not in st.session_state:
    credentials = to_dict(st.secrets["credentials"])   # ← そのまま

    authenticator = stauth.Authenticate(
        credentials, "salesdash", "salesdash_key", cookie_expiry_days=7
    )

    # ★ v0.4 以降の login ★
    fields = {
        "Form name": "ログイン",
        "Username": "ユーザー名",
        "Password": "パスワード",
        "Login": "ログイン"
    }
    name, auth_status, username = authenticator.login(
        fields=fields,       # ← ここが必須
        location="main"
    )

    if not auth_status:
        st.stop()

# ★★ ② ログアウト直後にフラグを消すなら ↓ を追加 ★★
authenticator.logout("ログアウト", "sidebar")
st.session_state.pop("auth_ok", None)     # ← 追加

# まずは分かっている別名をここに追加していく

# ========= 店舗名エイリアスの外部読込 =========
import os, json, pathlib
import streamlit as st

# YAML は任意（ファイルで YAML を使う場合のみ必要）
try:
    import yaml
except Exception:
    yaml = None

# TOML 読みは「任意」：外部 TOML ファイルを読むときだけ使う
# secrets.toml は streamlit が st.secrets で読んでくれるので tomllib は不要
try:
    import tomllib  # Python 3.11+
except Exception:
    tomllib = None


@st.cache_data(show_spinner=False)
def load_store_aliases() -> dict[str, str]:
    """
    店舗エイリアスを外部から読み込む。
    優先順位: st.secrets["store_aliases"] > 環境変数 STORE_ALIASES_FILE > 既定パス
    フォーマット: YAML / JSON / TOML（拡張子で自動判別）
    """
    # 1) secrets（本番でのパスワード同様の管理）
    if "store_aliases" in st.secrets:
        # secrets.toml の [store_aliases] セクションを dict に
        try:
            data = dict(st.secrets["store_aliases"])
            return {str(k): str(v) for k, v in data.items()}
        except Exception:
            pass

    # 2) ファイルパス指定（環境変数）
    path = os.environ.get("STORE_ALIASES_FILE")
    # 3) 既定の検索パス（存在する最初のファイルを採用）
    candidate_paths = [
        path,
        "config/store_aliases.yaml",
        "config/store_aliases.yml",
        "config/store_aliases.json",
        "config/store_aliases.toml",
        ".streamlit/store_aliases.toml",
    ]
    for p in candidate_paths:
        if not p:
            continue
        p = pathlib.Path(p)
        if not p.exists():
            continue
        try:
            ext = p.suffix.lower()
            with p.open("rb") as f:
                if ext in (".yml", ".yaml") and yaml is not None:
                    data = yaml.safe_load(f) or {}
                elif ext == ".json":
                    data = json.load(f) or {}
                elif ext == ".toml":
                    if tomllib is None:
                        data = {}  # tomllib が無ければスキップ
                    else:
                        t = tomllib.load(f) or {}
                        data = t.get("store_aliases", t)
                else:
                    data = {}
            if isinstance(data, dict):
                return {str(k): str(v) for k, v in data.items()}
        except Exception:
            # 壊れていても落とさずスキップ
            continue

    # 4) 見つからなければ空
    return {}


STORE_ALIASES: dict[str, str] = load_store_aliases()


import hashlib
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
def months_for_store_year(store: str, year: int,
                          sales_df: pd.DataFrame | None,
                          yj_df: pd.DataFrame | None,
                          pp_df: pd.DataFrame | None) -> list[int]:
    """店舗×年に対して、sales_df / yj_df / pp_df のいずれかに
    月データが存在する月を合集合で返す（昇順）。"""

    def _pick(df):
        if df is None or df.empty:
            return set()
        cols = df.columns
        # 必要カラムがそろっているデータだけ対象
        if not {"店舗名", "年", "月"}.issubset(cols):
            return set()
        sub = df[(df["店舗名"] == store) &
                 (pd.to_numeric(df["年"], errors="coerce") == int(year))]
        if sub.empty:
            return set()
        return set(pd.to_numeric(sub["月"], errors="coerce")
                   .dropna().astype(int).unique().tolist())

    months = _pick(sales_df) | _pick(yj_df) | _pick(pp_df)
    return sorted(m for m in months if 1 <= m <= 12)
def safe_year_month_for_sales(sel_year, sel_month, sales_df=None):
    """
    売上ファイルが未読込でも年・月を安全に返す。
    sales_df があれば、最新の年を優先し、月は選択値がなければ最大月。
    sales_df が無ければ、選択値→なければ今日の年・月。
    """
    try:
        if sales_df is not None and not sales_df.empty:
            latest = int(pd.to_numeric(sales_df["年"], errors="coerce").max())
            # 月の候補（その年に存在する月）
            m_opts = sorted(pd.to_numeric(
                sales_df.query("年 == @latest")["月"], errors="coerce"
            ).dropna().astype(int).unique().tolist())
            if not m_opts:
                return int(sel_year) if sel_year else datetime.now().year, \
                       int(sel_month) if sel_month else datetime.now().month
            # sel_month が候補にない/None の場合は最大月を採用
            mm = int(sel_month) if sel_month and int(sel_month) in m_opts else m_opts[-1]
            return latest, mm
    except Exception:
        pass

    # sales_df が無い or 取得に失敗 → セレクタ値 or 今日
    y = int(sel_year) if sel_year else datetime.now().year
    m = int(sel_month) if sel_month else datetime.now().month
    return y, m
def _blob(up):
    """UploadedFile -> (bytes, name)"""
    data = up.getvalue() if hasattr(up, "getvalue") else (up.seek(0) or up.read())
    if not isinstance(data, (bytes, bytearray)):
        data = b""
    return data, getattr(up, "name", "noname.xlsx")

@st.cache_data(show_spinner=False)
def parse_sales_cached(blob: bytes, name: str):
    """売上管理：1ファイル分をキャッシュ付で解析"""
    b = BytesIO(blob)
    df = read_sales_sheet(b)
    y, m = infer_year_month(df)
    store = _store_from_filename(name)

    b.seek(0); ltv  = parse_ltv(b)
    b.seek(0); card = parse_card_count(b)

    agg = {
        "総売上": pd.to_numeric(df["総売上"], errors="coerce").sum(),
        "総来院数": pd.to_numeric(df["総来院数"], errors="coerce").sum(),
        "保険新患": pd.to_numeric(df["保険新患"], errors="coerce").sum(),
        "自由新患": pd.to_numeric(df["自由新患"], errors="coerce").sum(),
    }
    return {"store":store, "year":y, "month":m, "agg":agg, "ltv":ltv, "card":card}

@st.cache_data(show_spinner=False)
def parse_yj_cached(blob: bytes, name: str):
    """予実：1ブック分をキャッシュ付で解析"""
    df = _yj_parse_with_sheet_progress(BytesIO(blob))
    if df is None or df.empty: 
        return pd.DataFrame()
    df = df.copy()
    df["元ファイル"] = name
    return df

@st.cache_data(show_spinner=False)
def parse_person_cached(blob: bytes, name: str):
    """個人生産性：1ブック分をキャッシュ付で解析"""
    df, _logs = parse_person_book(BytesIO(blob), name)
    return df if df is not None else pd.DataFrame()

def _parallel_map(func, files):
    """並列解析（ファイル数が多い時に有効）"""
    with ThreadPoolExecutor(max_workers=min(8, len(files))) as ex:
        futs = [ex.submit(func, *_blob(up)) for up in files]
        for fu in as_completed(futs):
            yield fu.result()
def _norm_store_core(s: str) -> str:
    """ファイル名/表記の揺れを吸収して店名のコアを取り出す"""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.strip()
    s = re.sub(r"\.xlsx$", "", s, flags=re.I)    # 拡張子
    s = re.sub(r"^\d+[._\-\s]*", "", s)          # 先頭の番号
    s = re.sub(r"\s+", "", s)                    # 空白
    s = re.sub(r"\(.*?\)|（.*?）", "", s)         # カッコ内
    s = re.sub(r"\d{1,2}月.*$", "", s)           # 「1月…」など
    # 余計な語を除去
    s = s.replace("株式会社","").replace("有限会社","")
    s = s.replace("店","").replace("院","").replace("支店","").replace("本店","")
    s = re.sub(r"(イーグル|クレーン|EAGLE|ｅａｇｌｅ|ＹＴ|YT|ＹＢ|YB)", "", s, flags=re.I)
    return s

def canonical_store_name(raw: str) -> str:
    core = _norm_store_core(raw)
    if core in STORE_ALIASES:
        return STORE_ALIASES[core]
    # 未登録は「イーグル＋コア」に寄せる
    return ("イーグル" + core) if core else "イーグル不明"

def unify_store_names(df: pd.DataFrame, col="店舗名") -> pd.DataFrame:
    if col in df.columns:
        out = df.copy()
        out[col] = out[col].apply(canonical_store_name)
        return out
    return df



# ------------------ ユーティリティ ------------------

def _num(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s)
    s = re.sub(r"[▲△−\-]", "-", s)
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s) if s != "" else None
    except Exception:
        return None
def _metric_card(col, label, cur, prv, unit=""):
    """st.metric 用のラッパー（前年比%を自動計算/安全化）。"""
    import math
    cur_txt = "-" if cur is None or (isinstance(cur, float) and math.isnan(cur)) else f"{int(round(cur)):,}{unit}"
    delta_txt = ""
    if (prv is not None) and (not (isinstance(prv, float) and math.isnan(prv))) and prv != 0:
        delta = (cur - prv) / prv * 100 if (cur is not None and not (isinstance(cur, float) and math.isnan(cur))) else 0.0
        delta_txt = f"{delta:+.1f}% vs 前年同月"
    with col:
        st.metric(label, cur_txt, delta_txt)

def fmt_comma_int(df: pd.DataFrame, cols):
    """金額/件数列を 四捨五入→整数→カンマ文字列 に変換（NAは空文字）"""
    cols = [c for c in cols if c in df.columns]
    if not cols:
        return df
    for c in cols:
        s = pd.to_numeric(df[c], errors="coerce").round(0)
        df[c] = s.apply(lambda v: "" if pd.isna(v) else f"{int(v):,}")
    return df
def fmt_percent(df: pd.DataFrame, cols):
    """比率(0.0〜1.0)でも百分率(0〜100)でも受け取り、xx.x% 文字列に統一"""
    cols = [c for c in cols if c in df.columns]
    if not cols:
        return df
    for c in cols:
        s = pd.to_numeric(df[c], errors="coerce")
        # 多数が 0〜1 に収まる場合は比率とみなして 100 倍
        with pd.option_context('mode.use_inf_as_na', True):
            non_na = s.dropna()
        if not non_na.empty and (non_na.between(0, 1).mean() >= 0.6):
            s = s * 100
        df[c] = s.round(1).apply(lambda v: "" if pd.isna(v) else f"{v:.1f}")
    return df
def sty(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
    """% 以外は小数点なし、年はカンマなし"""
    fmts: dict[str, str] = {}
    for c in df.columns:
        if c.endswith("%"):
            fmts[c] = "{:+.1f}"  # 増減率は ±1 桁
        elif pd.api.types.is_numeric_dtype(df[c]):
            fmts[c] = "{:,.0f}"   # 整数(千区切り) 小数無し
        if c == "年":
            fmts[c] = "{:d}"    # カンマなし
    return df.style.format(fmts)
def _has_cols(df: pd.DataFrame | None, cols: list[str]) -> bool:
    return isinstance(df, pd.DataFrame) and (not df.empty) and set(cols).issubset(df.columns)

# ------------------ LTV / カルテ（Excel由来） ------------------
def parse_ltv(file_bytes: io.BytesIO):
    try:
        file_bytes.seek(0)
        df = pd.read_excel(file_bytes, sheet_name="店舗分析", header=None, engine="openpyxl")
    except Exception:
        return None
    tgt_col = None
    for r in range(0, min(30, df.shape[0])):
        for c in range(df.shape[1]):
            if re.search(r"(今月|当月)実績", str(df.iat[r, c])):
                tgt_col = c; break
        if tgt_col is not None: break
    row_idx = None
    for r in range(df.shape[0]):
        row_text = " ".join(str(x) for x in df.iloc[r, :5].tolist())
        if "LTV" in row_text:
            row_idx = r; break
    if row_idx is None or tgt_col is None:
        return None
    return _num(df.iat[row_idx, tgt_col])

def parse_card_count(file_bytes: io.BytesIO):
    try:
        file_bytes.seek(0)
        df = pd.read_excel(file_bytes, sheet_name="店舗分析", header=None, engine="openpyxl")
    except Exception:
        return None
    tgt_col = None
    for r in range(0, min(30, df.shape[0])):
        for c in range(df.shape[1]):
            if re.search(r"(今月|当月)実績", str(df.iat[r, c])):
                tgt_col = c; break
        if tgt_col is not None: break
    row_idx = None
    for r in range(df.shape[0]):
        row_text = " ".join(str(x) for x in df.iloc[r, :5].tolist())
        if re.search(r"カルテ.*枚", row_text):
            row_idx = r; break
    if row_idx is None or tgt_col is None:
        return None
    return _num(df.iat[row_idx, tgt_col])

# ------------------ 売上管理表 読み取り ------------------
def _detect_header_row(df_raw: pd.DataFrame) -> int:
    for i in range(min(40, len(df_raw))):
        if df_raw.iloc[i].astype(str).str.contains("日付").any():
            return i
    return 0

def read_sales_sheet(file_bytes: io.BytesIO) -> pd.DataFrame:
    file_bytes.seek(0)
    raw = pd.read_excel(file_bytes, sheet_name="売上管理", header=None, engine="openpyxl")
    header_row = _detect_header_row(raw)
    header = raw.iloc[header_row].astype(str).tolist()

    date_idx = next((j for j, v in enumerate(header) if "日付" in str(v)), 0)
    data_rows = raw.iloc[header_row+1:].copy()

    def find_right_neighbor_idx(hdr, pat):
        s = pd.Series(hdr).astype(str)
        hit = s[s.str.contains(pat, na=False)]
        if hit.empty:
            return None
        base_idx = int(hit.index[0]); neighbor = base_idx + 1
        return neighbor if neighbor < len(hdr) else None

    idx_ins = find_right_neighbor_idx(header, r"保険診療")
    idx_sp  = find_right_neighbor_idx(header, r"自由診療")
    if idx_ins is None: idx_ins = 4
    if idx_sp  is None: idx_sp  = 7

    valid_mask = pd.to_datetime(data_rows.iloc[:, date_idx], errors="coerce").notna()
    ins_abs = pd.to_numeric(data_rows.iloc[:, idx_ins], errors="coerce").where(valid_mask, 0).fillna(0)
    sp_abs  = pd.to_numeric(data_rows.iloc[:, idx_sp ], errors="coerce").where(valid_mask, 0).fillna(0)

    file_bytes.seek(0)
    df = pd.read_excel(file_bytes, sheet_name="売上管理", header=header_row, engine="openpyxl")
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
    df = df.dropna(subset=["日付"]).reset_index(drop=True)

    def _find_col(patterns):
        for c in df.columns:
            for p in patterns:
                if re.search(p, str(c)):
                    return c
        return None
    if "総売上" not in df.columns:
        alt_sales = _find_col([r"総売上", r"売上合計", r"合計売上", r"売上\s*計"])
        if alt_sales is not None:
            df["総売上"] = pd.to_numeric(df[alt_sales], errors="coerce").fillna(0)
    if "総来院数" not in df.columns:
        alt_vis = _find_col([r"総来院", r"来院数", r"患者数"])
        if alt_vis is not None:
            df["総来院数"] = pd.to_numeric(df[alt_vis], errors="coerce").fillna(0)

    for col in ("総売上", "総来院数"):
        if col in df.columns:
            df[col] = (
                df[col].astype(str).str.replace(r"[▲△▲−-]", "-", regex=True)
                                 .str.replace(r"[^0-9\.-]", "", regex=True)
                                 .replace("", "0")
            )
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["保険新患"] = ins_abs.reset_index(drop=True).iloc[:len(df)].to_numpy()
    df["自由新患"] = sp_abs.reset_index(drop=True).iloc[:len(df)].to_numpy()

    cols = [c for c in ["日付","総売上","総来院数","保険新患","自由新患"] if c in df.columns]
    return df[cols]
# ───── 患者分析シート ─────

def parse_patient_analysis(f, add_msg):
    """患者分析シートを抽出。無い/欠落時は 0 データで返す"""
    # 決め打ちカテゴリ
    C_GENDER = ["男性", "女性"]
    C_REASON = ["チラシ", "紹介", "看板", "ネット", "その他"]
    C_AGE    = ["10代未満", "10代", "20代", "30代", "40代", "50代", "60代", "70代", "80代", "90歳以上"]

    zero = lambda cats: pd.DataFrame({"カテゴリ": cats, "件数": [0]*len(cats)})

    try:
        xls = pd.ExcelFile(f, engine="openpyxl")
        if "患者分析" not in xls.sheet_names:
            raise ValueError("シートなし")
        sheet = xls.parse("患者分析", header=None)
    except Exception:
        add_msg(f"{f.name}: 患者分析シートが見つかりません - 0 件として処理します")
        return zero(C_GENDER), zero(C_REASON), zero(C_AGE)

    def grab(keyword: str, rng: slice | None, cats: list[str]):
        mask = sheet.apply(lambda r: r.astype(str).str.contains(keyword).any(), axis=1)
        if not mask.any():
            return zero(cats)
        r = mask.idxmax()
        # データ行が足りない場合は 0
        if r + 2 >= len(sheet):
            add_msg(f"{f.name}: 患者分析シートが見つかりません - 0 件として処理します")
            return zero(cats)
        header = sheet.iloc[r + 1]
        vals   = sheet.iloc[r + 2]
        if rng is not None:
            header = header.iloc[rng]
            vals   = vals.iloc[rng]
        header = header.dropna()
        if header.empty:
            return zero(cats)
        data = pd.to_numeric(vals[header.index], errors="coerce").fillna(0)
        return pd.DataFrame({"カテゴリ": header.values, "件数": data.values})

    gender = grab("男女比率",  slice(0, 2),  C_GENDER)  # A:B
    reason = grab("来院動機", slice(5, 10), C_REASON)  # F:J
    age    = grab("年齢比率", None,        C_AGE)
    return gender, reason, age
# 患者分析プロット
def plot_pivot(df_src, store, latest, title):
    # ★列が無い/空なら何も描かず返す（未読込時の防弾）
    if not _has_cols(df_src, ["店舗名", "年", "カテゴリ", "件数"]):
        st.info(f"{title}のデータがありません。")
        return

    # ブールインデックスで抽出（queryは使わない）
    mask = (
        (df_src["店舗名"].astype(str) == str(store)) &
        (pd.to_numeric(df_src["年"], errors="coerce") == int(latest))
    )
    df = (df_src.loc[mask, ["カテゴリ", "件数"]]
                 .groupby("カテゴリ", as_index=False)["件数"].sum())

    if df.empty:
        st.info(f"{store}（{latest}年）の{title}データがありません。")
        return

    df["割合%"] = (df["件数"] / df["件数"].sum() * 100).round(1)
    st.altair_chart(
        alt.Chart(df).mark_bar().encode(
            y=alt.Y("カテゴリ:N", sort="-x"),
            x="件数:Q",
            tooltip=["カテゴリ", "件数", "割合%"],
        ).properties(width=350, height=250, title=title),
        use_container_width=True,
    )
    with st.expander(f"📄 {title} 明細"):
        st.dataframe(df, use_container_width=True)

# ───── 任意カテゴリ：前年比較プロット ─────
# ───── 任意カテゴリ：前年比較プロット ─────
def plot_cat_yoy(df_src, store, latest, prev, title):
    # ★列が無い/空なら何も描かず返す
    if not _has_cols(df_src, ["店舗名", "年", "カテゴリ", "件数"]):
        st.info(f"{title}のデータがありません。")
        return

    # 今年/前年をブールインデックスで抽出
    cur_mask = ((df_src["店舗名"].astype(str) == str(store)) &
                (pd.to_numeric(df_src["年"], errors="coerce") == int(latest)))
    prv_mask = ((df_src["店舗名"].astype(str) == str(store)) &
                (pd.to_numeric(df_src["年"], errors="coerce") == int(prev)))

    cur = (df_src.loc[cur_mask, ["カテゴリ", "件数"]]
                 .groupby("カテゴリ", as_index=False)["件数"].sum()
                 .rename(columns={"件数": "今年"}))
    old = (df_src.loc[prv_mask, ["カテゴリ", "件数"]]
                 .groupby("カテゴリ", as_index=False)["件数"].sum()
                 .rename(columns={"件数": "前年"}))

    comp = pd.merge(cur, old, on="カテゴリ", how="outer").fillna(0)

    # ---------- グラフ ----------
    comp_melt = (comp.rename(columns={"前年": str(prev), "今年": str(latest)})
                       .melt(id_vars="カテゴリ",
                             value_vars=[str(prev), str(latest)],
                             var_name="年度", value_name="件数"))

    chart = (
        alt.Chart(comp_melt)
           .mark_bar()
           .encode(
               x=alt.X("カテゴリ:N", sort="-y", title=title),
               y="件数:Q",
               xOffset=alt.XOffset("年度:N",
                                   scale=alt.Scale(domain=[str(prev), str(latest)])),
               color=alt.Color("年度:N",
                               scale=alt.Scale(domain=[str(prev), str(latest)],
                                               range=["#4e79a7", "#a0cbe8"])),
               tooltip=["年度", "カテゴリ", "件数"],
           )
           .properties(width=400, height=300,
                       title=f"{store} {title} ({prev} vs {latest})")
    )
    st.altair_chart(chart, use_container_width=True)

    # ---------- 増減テーブル ----------
    diff_tbl = (comp.set_index("カテゴリ")
                     .apply(pd.to_numeric, errors="coerce")
                     .fillna(0))
    diff_tbl["増減差"]  = diff_tbl["今年"] - diff_tbl["前年"]
    diff_tbl["増減率%"] = np.where(
        diff_tbl["前年"] == 0, np.nan,
        (diff_tbl["増減差"] / diff_tbl["前年"] * 100).round(1)
    )
    with st.expander(f"📄 {title} 増減明細"):
        st.dataframe(sty(diff_tbl.reset_index()), use_container_width=True)


def infer_year_month(df: pd.DataFrame) -> Tuple[int, int]:
    day = pd.to_datetime(df["日付"], errors="coerce").dropna()
    if day.empty:
        now = pd.Timestamp.today()
        return int(now.year), int(now.month)
    return int(day.dt.year.mode()[0]), int(day.dt.month.mode()[0])

def _store_from_filename(name: str) -> str:
    base = re.sub(r"\.xlsx$", "", name)
    base = re.sub(r"^\d+[\.\s＿-]*", "", base)
    base = re.sub(r"\s*\d{1,2}月.*$", "", base)
    return canonical_store_name(base)   # ← ここを追加

# 置き換え：@st.cache_data を付け、オプションを追加
@st.cache_data(show_spinner=False)
def load(files, *, read_ltv=True, read_card=True, read_patient=True):
    total = len(files) if files else 0
    bar = st.progress(0, text="📥 売上管理表を読み取り準備…") if total else None

    sales, reasons, genders, ages, ltvs, cards = [], [], [], [], [], []
    msgs = []
    for i, up in enumerate(files, start=1):
        if bar:
            bar.progress(int(i*100/total), text=f"📖 {getattr(up,'name','ファイル')} を読み取り中…（{i}/{total}）")
        name = getattr(up, "name", "売上管理表.xlsx")
        try:
            data = up.getvalue() if hasattr(up, "getvalue") else (up.seek(0) or up.read())
            file_bytes = BytesIO(data) if isinstance(data, (bytes, bytearray)) else up
            setattr(file_bytes, "name", name)

            # 売上・来院・新患（必須）
            df = read_sales_sheet(file_bytes)
            y, m = infer_year_month(df)
            store = _store_from_filename(name)
            agg = {
                "総売上":   pd.to_numeric(df["総売上"], errors="coerce").sum(),
                "総来院数": pd.to_numeric(df["総来院数"], errors="coerce").sum(),
                "保険新患": pd.to_numeric(df["保険新患"], errors="coerce").sum(),
                "自由新患": pd.to_numeric(df["自由新患"], errors="coerce").sum(),
            }
            sales.append({"店舗名": store, "年": y, "月": m, **agg})

            # LTV / カルテ（オプション）
            if read_ltv:
                ltv = parse_ltv(file_bytes)
                if ltv is not None:
                    ltvs.append({"店舗名": store, "年": y, "月": m, "LTV": ltv})
            if read_card:
                card = parse_card_count(file_bytes)
                if card is not None:
                    cards.append({"店舗名": store, "年": y, "月": m, "カルテ枚数": card})

        except Exception as e:
            msgs.append(f"{name}: 読み取り失敗（{e}）")

        # 患者分析（オプション）
        if read_patient:
            def _add_msg(msg): msgs.append(str(msg))
            try:
                file_bytes.seek(0)
                gdf, rdf, adf = parse_patient_analysis(file_bytes, _add_msg)
                for _df, bucket in [(rdf, reasons), (gdf, genders), (adf, ages)]:
                    if _df is not None and not _df.empty:
                        d = _df.copy()
                        d["店舗名"] = store; d["年"] = y; d["月"] = m
                        bucket.append(d)
            except Exception as e:
                msgs.append(f"{name}: 患者分析 読み取り失敗（{e}）")

    if bar:
        bar.progress(100, text=f"✅ 売上管理表 {total} 件 取り込み完了")
        time.sleep(0.2); bar.empty()

    def _to_df(lst):
        if not lst: return pd.DataFrame()
        first = lst[0]
        if isinstance(first, (pd.DataFrame, pd.Series)):
            return pd.concat(lst, ignore_index=True)
        return pd.DataFrame(lst)

    return _to_df(sales), _to_df(reasons), _to_df(genders), _to_df(ages), pd.DataFrame(ltvs), pd.DataFrame(cards), msgs


# ------------------ 予実管理（シート進捗つき） ------------------
def _yj_parse_with_sheet_progress(file_like):
    import numpy as np, re, pandas as pd
    from openpyxl import load_workbook
    from contextlib import suppress

    def to_num(x):
        if x is None:
            return float("nan")
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x)
        s = re.sub(r"[▲△−\-]", "-", s)
        s = re.sub(r"[^0-9\.\-]", "", s)
        try:
            return float(s) if s != "" else float("nan")
        except Exception:
            return float("nan")

    exp_keys = [
        "広告宣伝費","荷造運賃","給料手当","法定福利費","厚生費","減価償却費","負債料","修繕費","保険料","賃借料",
        "事務用品費","消耗品費","水道光熱費","旅費交通費","手数料","租税公課","通信費","諸会費",
        "新聞図書費","地代家賃","燃料費","リース料","雑費","駐車場負担金","駐車場負担金（患者）","柔道整会負担金","研修費"
    ]
    sale_keys = ["純売上", "純売上高", "売上合計", "総売上", "売上高"]

    wb = load_workbook(file_like, read_only=True, data_only=True)
    sheet_names = [s for s in wb.sheetnames if all(x not in s for x in ["一覧","合計"])]

    rows = []
    try:
        st_status = st.status("📘 予実ブックを解析中…", expanded=False)
    except Exception:
        st_status = None

    for i, sheet in enumerate(sheet_names, start=1):
        if st_status:
            with suppress(Exception):
                st_status.update(label=f"[{i}/{len(sheet_names)}] 『{sheet}』を読み取り中…")

        ws = wb[sheet]
        MAX_R, MAX_C = 180, 200

        header_row = item_col = year = None
        for r in range(1, min(MAX_R, ws.max_row)+1):
            for c in range(1, min(MAX_C, ws.max_column)+1):
                if ws.cell(r, c).value == "科目":
                    header_row, item_col = r, c
                    for c2 in range(c+1, min(c+60, ws.max_column)+1):
                        hv = str(ws.cell(r, c2).value or "")
                        m = re.search(r"(\d{4})目標", hv)
                        if m:
                            year = int(m.group(1)); break
                    break
            if header_row is not None: break
        if header_row is None or item_col is None:
            continue

        def _left_group_label(col):
            for cc in range(col, max(item_col, col-10), -1):
                s = ws.cell(header_row, cc).value
                if s is not None and str(s).strip() != "":
                    return str(s)
            return ""

        tgt_cols, act_cols = {}, {}
        for c in range(item_col+1, min(item_col+80, ws.max_column)+1):
            sub = str(ws.cell(header_row+1, c).value or "")
            m = re.fullmatch(r"(\d{1,2})月", sub)
            if not m:
                continue
            top_label = _left_group_label(c)
            if "実績" in top_label:
                mth = int(m.group(1))
                act_cols[mth] = c
                tgt_cols[mth] = c - 2

        if not act_cols:
            continue

        sale_row = None
        exp_rows = []
        for r in range(header_row+1, min(header_row+100, ws.max_row)+1):
            label = str(ws.cell(r, item_col).value or "")
            if any(k in label for k in sale_keys): sale_row = r
            if any(k in label for k in exp_keys): exp_rows.append(r)

        store = re.sub(r"^\d+\s*", "", sheet).strip()
        months = sorted(act_cols.keys())
        max_act = months[-1] if months else None

        top_goal = None
        for r in range(1, 40):
            for c in range(1, 80):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "売上目標" in v:
                    n = to_num(ws.cell(r, c+1).value)
                    if n == n: top_goal = n; break
            if top_goal is not None: break

        for mth in months:
            rec = {"店舗名": store, "年": year, "月": mth}
            if sale_row is not None:
                if mth in tgt_cols: rec["売上目標"] = to_num(ws.cell(sale_row, tgt_cols[mth]).value)
                if mth in act_cols: rec["売上実績"] = to_num(ws.cell(sale_row, act_cols[mth]).value)
            def _nansum_cells(ws, rows, col):
                vals = [to_num(ws.cell(r, col).value) for r in rows]
                return float(np.nansum(vals))  # NaN を 0 として合計

            if exp_rows:
                if mth in tgt_cols: rec["経費目標"] = _nansum_cells(ws, exp_rows, tgt_cols[mth])
                if mth in act_cols: rec["経費実績"] = _nansum_cells(ws, exp_rows, act_cols[mth])
            
            rows.append(rec)

        if rows and top_goal is not None and max_act is not None:
            for rec in reversed(rows):
                if rec.get("店舗名")==store and rec.get("年")==year and rec.get("月")==max_act:
                    rec["売上目標"] = top_goal
                    break

    out = pd.DataFrame(rows)
    if st_status:
        with suppress(Exception):
            st_status.update(label="✅ 予実ブック解析 完了", state="complete")

    if not out.empty:
        import numpy as np
        out["粗利目標"] = out.get("売上目標", np.nan) - out.get("経費目標", np.nan)
        out["粗利実績"] = out.get("売上実績", np.nan) - out.get("経費実績", np.nan)
        out["粗利率目標"] = np.where(out.get("売上目標", np.nan).fillna(0)==0, np.nan, out["粗利目標"]/out["売上目標"]*100)
        out["粗利率実績"] = np.where(out.get("売上実績", np.nan).fillna(0)==0, np.nan, out["粗利実績"]/out["売上実績"]*100)
    return out

def parse_yj_workbook(file_like):
    return _yj_parse_with_sheet_progress(file_like)

@st.cache_data(show_spinner=False)
def load_yj(files):
    if not files:
        return pd.DataFrame()
    dfs = []
    total = len(files)
    bar = st.progress(0, text="📥 予実管理ファイルを読み取り準備…")
    for i, up in enumerate(files, start=1):
        name = getattr(up, "name", "予実ブック.xlsx")
        bar.progress(int(i*100/total), text=f"📖 {name} を読み取り中…（{i}/{total}）")
        try:
            data = up.getvalue() if hasattr(up, "getvalue") else (up.seek(0) or up.read())
            buf = BytesIO(data) if isinstance(data, (bytes, bytearray)) else up
            df = _yj_parse_with_sheet_progress(buf)
            if df is not None and not df.empty:
                df = df.copy(); df["元ファイル"] = name
                dfs.append(df)
        except Exception as e:
            st.warning(f"予実ファイルの読込に失敗: {name} … {e}")
    if not dfs:
        bar.empty(); return pd.DataFrame()
    bar.progress(100, text=f"✅ 予実管理ファイル {total} 件 取り込み完了")
    time.sleep(0.2); bar.empty()
    yj = pd.concat(dfs, ignore_index=True)
    yj = yj.sort_values(["店舗名","年","月"]).drop_duplicates(subset=["店舗名","年","月"], keep="last")
    return yj

# ------------------ 個人生産性（新実装） ------------------
def _guess_year_month_from_sheet(df: pd.DataFrame, sheet_name: str, file_name: str):
    """
    年月の推定を強化：
      - まずタブ名 'YYMM' を 20YY/MM と解釈（例: 2506→2025/6）
      - 次に '20YY' + 区切り + 'MM'（例: 2025-06, 2025_6, 202506）
      - シート上部のテキストは「年」「月」の文脈がある組合せのみ採用
      - ファイル名からの保険（202506 / 2025-06 / YYMM）にも対応
    """
    sname = str(sheet_name).strip()
    fname = str(file_name).strip()

    # 1) タブ名が YYMM（4桁のみ）なら 2000+YY / MM
    m = re.fullmatch(r"(\d{2})([01]\d)", sname)
    if m:
        yy, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return 2000 + yy, mm

    # 2) タブ名: 20YY + 区切り? + MM
    m = re.search(r"(20\d{2})\D{0,3}([01]?\d)\b", sname)
    if m:
        y, mth = int(m.group(1)), int(m.group(2))
        if 1 <= mth <= 12:
            return y, mth

    # 3) シート上部（A1～の数行×数列）から、年・月の“組合せ”で拾う
    def _top_blob(df):
        top = df.iloc[:8, :10].astype(str).fillna("")
        return " ".join(top.values.ravel())

    blob = _top_blob(df)

    # 2025年6月 / 6月 2025年 など
    m = re.search(r"(20\d{2})\s*年\D{0,3}([01]?\d)\s*月", blob)
    if m:
        y, mth = int(m.group(1)), int(m.group(2))
        if 1 <= mth <= 12:
            return y, mth

    m = re.search(r"([01]?\d)\s*月\D{0,5}(20\d{2})", blob)
    if m:
        mth, y = int(m.group(1)), int(m.group(2))
        if 1 <= mth <= 12:
            return y, mth

    # 月だけ見つかったら、年はファイル名/タブ名の 20YY を採用
    m = re.search(r"([01]?\d)\s*月", blob)
    if m:
        mth = int(m.group(1))
        y = None
        m_y = re.search(r"(20\d{2})", sname) or re.search(r"(20\d{2})", fname)
        if m_y:
            y = int(m_y.group(1))
        if y and 1 <= mth <= 12:
            return y, mth

    # 4) ファイル名からの保険: 202506 / 2025-06 / YYMM
    m = re.search(r"(20\d{2})\D{0,3}([01]?\d)\b", fname)
    if m:
        y, mth = int(m.group(1)), int(m.group(2))
        if 1 <= mth <= 12:
            return y, mth
    m = re.search(r"(?<!\d)(\d{2})([01]\d)(?!\d)", fname)
    if m:
        yy, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return 2000 + yy, mm

    return None, None

    """シート上部・シート名・ファイル名から 年・月 を推定"""
    year = None; month = None
    # 1) シート上部
    top = df.iloc[:6, :8].astype(str).fillna("")
    blob = " ".join(top.values.ravel())
    m = re.search(r"(\d{1,2})\s*月", blob)
    if m: month = int(m.group(1))
    y = re.search(r"(20\d{2})", blob)
    if y: year = int(y.group(1))
    # 2) シート名
    if month is None:
        m2 = re.search(r"(\d{1,2})\s*月", sheet_name)
        if m2: month = int(m2.group(1))
    if year is None:
        y2 = re.search(r"(20\d{2})", sheet_name)
        if y2: year = int(y2.group(1))
    # 3) ファイル名
    if year is None:
        y3 = re.search(r"(20\d{2})", file_name)
        if y3: year = int(y3.group(1))
    return year, month

def parse_person_book(file_like, filename: str):
    """各月シートから 院/担当者 の『稼働数・目標値/日・目標値/月・進捗・達成率』を抽出"""
    logs = []
    xls = pd.ExcelFile(file_like, engine="openpyxl")
    store = _store_from_filename(filename)
    recs = []

    for sh in xls.sheet_names:
        df = xls.parse(sh, header=None)
        if df.empty:
            logs.append(f"{filename}:{sh} … 空シート")
            continue

        y, m = _guess_year_month_from_sheet(df, sh, filename)
        if not (y and m and 2000 <= int(y) <= 2100 and 1 <= int(m) <= 12):
            logs.append(f"{filename}:{sh} … 年月が判定できず")
            continue

        firstcol = df.iloc[:, 0].astype(str)
        idx_kado = firstcol[firstcol.str.contains("稼働数", na=False)].index
        if len(idx_kado) == 0:
            logs.append(f"{filename}:{sh} … 『稼働数』行が見つからない")
            continue
        header_row = int(idx_kado[0] - 1)

        # --- 担当（B列以降）。'0' / '０' / '0.0' が来たら以降を無視 ---
        raw_cells = df.iloc[header_row, 1:].tolist()
        staff = []
        for cell in raw_cells:
            s = "" if cell is None else str(cell).strip()
            # '0', '０', '0.0', '0.', '0,0' などを 0 とみなす
            s_norm = s.replace("０", "0").replace(",", "")
            if s_norm == "" or re.fullmatch(r"0(?:\.0+)?", s_norm):
                break
            if s.lower() == "nan":
                break
            staff.append(s)
        if not staff:
            logs.append(f"{filename}:{sh} … 担当者見出し（B列以降）が空")
            continue
        staff_len = len(staff)

        def find_row(pat):
            hit = firstcol[firstcol.str.contains(pat, na=False)]
            return int(hit.index[0]) if len(hit) else None

        rows = {
            "稼働数": find_row(r"稼働数"),
            "目標値_日": find_row(r"目標値\s*[\/／]?\s*日"),
            "目標値_月": find_row(r"目標値\s*[\/／]?\s*月"),
            "進捗":   find_row(r"進捗"),
            "達成率": find_row(r"達成率"),
        }
        if any(v is None for v in rows.values()):
            logs.append(f"{filename}:{sh} … 必須行（稼働数/目標値/日/目標値/月/進捗/達成率）のどれかが不足")
            continue

        # 各行も 'staff_len' 列ぶんだけ取得（0 以降は切り捨て）
        def row_vals(ridx):
            vals = df.iloc[ridx, 1:1+staff_len].tolist()
            return [_num(v) for v in vals]

        vals = {k: row_vals(v) for k, v in rows.items()}

        for j, name in enumerate(staff):
            recs.append({
                "店舗名": store,
                "年": int(y), "月": int(m),
                "担当": name,                 # 「院」も含む
                "稼働数": vals["稼働数"][j] if j < len(vals["稼働数"]) else None,
                "目標値_日": vals["目標値_日"][j] if j < len(vals["目標値_日"]) else None,
                "目標値_月": vals["目標値_月"][j] if j < len(vals["目標値_月"]) else None,
                "進捗": vals["進捗"][j] if j < len(vals["進捗"]) else None,
                "達成率": vals["達成率"][j] if j < len(vals["達成率"]) else None,
            })
        logs.append(f"{filename}:{sh} … {staff_len}名 読み取り（{y}/{m}）")

    return pd.DataFrame(recs), logs

  

@st.cache_data(show_spinner=False)
def load_person_productivity(files):
    """個人生産性ファイル群を統合"""
    if not files:
        return pd.DataFrame(), []
    dfs, logs = [], []
    total = len(files)
    bar = st.progress(0, text="📥 個人生産性ファイルを読み取り準備…")
    for i, up in enumerate(files, start=1):
        name = getattr(up, "name", "person.xlsx")
        bar.progress(int(i*100/total), text=f"📖 {name} を解析中…（{i}/{total}）")
        try:
            data = up.getvalue() if hasattr(up, "getvalue") else (up.seek(0) or up.read())
            buf = BytesIO(data) if isinstance(data, (bytes, bytearray)) else up
            df_one, lg = parse_person_book(buf, name)
            logs.extend(lg)
            if not df_one.empty:
                dfs.append(df_one)
        except Exception as e:
            logs.append(f"{name}: 読み取りエラー … {e}")
    bar.progress(100, text="✅ 個人生産性 取り込み完了")
    time.sleep(0.2); bar.empty()
    return (pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()), logs


# ------------------ UI ------------------
st.title("📊 売上ダッシュボード")

# 1) 売上管理ファイル
st.header("📂 売上管理ファイル（店舗ごと・複数可）")
files_sales = st.file_uploader("📂 売上管理表（例: 01.店舗名 1月 売上管理表.xlsx）", type=["xlsx"], accept_multiple_files=True, key="sales_files")
# 追加：読み取りオプション（速度優先）
with st.expander("⚙️ 読み取りオプション（速度優先）", expanded=False):
    c1, c2, c3 = st.columns(3)
    with c1:
        opt_ltv = st.checkbox("LTV を読む", value=True)
    with c2:
        opt_card = st.checkbox("カルテ枚数を読む", value=True)
    with c3:
        opt_patient = st.checkbox("患者分析を読む", value=True)

sales_df = reason_df = gender_df = age_df = pd.DataFrame()
ltv_df = card_df = pd.DataFrame()
msgs = []
if files_sales:
    sales_df, reason_df, gender_df, age_df, ltv_df, card_df, msgs = load(
        files_sales,
        read_ltv=opt_ltv, read_card=opt_card, read_patient=opt_patient
    )
    # ③ 店名統一（売上管理）
    if not sales_df.empty:
        sales_df = unify_store_names(sales_df)
    if not ltv_df.empty:
        ltv_df = unify_store_names(ltv_df)
    if not card_df.empty:
        card_df = unify_store_names(card_df)
    if not reason_df.empty:
        reason_df = unify_store_names(reason_df)
    if not gender_df.empty:
        gender_df = unify_store_names(gender_df)
    if not age_df.empty:
        age_df = unify_store_names(age_df)
    if msgs:
        with st.expander("⚠️ 読み取りメッセージ"):
            for m in msgs:
                st.write(m)

# 2) 📊 全店舗サマリー（最新年度 vs 前年）
st.subheader("📊 全店舗サマリー（最新年度 vs 前年）")
if not sales_df.empty:
    monthly = (
        sales_df.groupby(["店舗名", "年", "月"], as_index=False)[
            ["総売上", "総来院数", "保険新患", "自由新患"]
        ].sum()
    )
    latest = int(monthly["年"].max()); prev = latest - 1
    cur = monthly[monthly["年"] == latest]
    old = monthly[monthly["年"] == prev]
    comp = pd.merge(cur, old, on=["店舗名","月"], how="left", suffixes=("_今年","_前年"))

    # LTV
    if not ltv_df.empty:
        ltv_cur = ltv_df[ltv_df["年"] == latest].rename(columns={"LTV":"LTV_今年"})
        ltv_old = ltv_df[ltv_df["年"] == prev].rename(columns={"LTV":"LTV_前年"})
        ltv_c = pd.merge(ltv_cur, ltv_old, on=["店舗名","月"], how="left")[["店舗名","月","LTV_今年","LTV_前年"]]
        comp = pd.merge(comp, ltv_c, on=["店舗名","月"], how="left")

    # カルテ枚数
    if not card_df.empty:
        card_cur = card_df[card_df["年"] == latest].rename(columns={"カルテ枚数":"カルテ枚数_今年"})
        card_old = card_df[card_df["年"] == prev].rename(columns={"カルテ枚数":"カルテ枚数_前年"})
        card_c = pd.merge(card_cur, card_old, on=["店舗名","月"], how="left")[["店舗名","月","カルテ枚数_今年","カルテ枚数_前年"]]
        comp = pd.merge(comp, card_c, on=["店舗名","月"], how="left")

    # 増減率（小数1桁）
    def _rate(a, b):
        den = b.replace(0, pd.NA) if isinstance(b, pd.Series) else (pd.NA if b==0 else b)
        return ((a - b) / den * 100).astype("Float64").round(1) if isinstance(den, pd.Series) else None

    comp["総売上増減率%"]   = _rate(comp["総売上_今年"], comp["総売上_前年"])
    comp["総来院数増減率%"] = _rate(comp["総来院数_今年"], comp["総来院数_前年"])
    comp["保険新患増減率%"] = _rate(comp["保険新患_今年"], comp["保険新患_前年"])
    comp["自由新患増減率%"] = _rate(comp["自由新患_今年"], comp["自由新患_前年"])
    if "LTV_今年" in comp.columns and "LTV_前年" in comp.columns:
        comp["LTV増減率%"] = _rate(comp["LTV_今年"], comp["LTV_前年"])
    if "カルテ枚数_今年" in comp.columns and "カルテ枚数_前年" in comp.columns:
        comp["カルテ枚数増減率%"] = _rate(comp["カルテ枚数_今年"], comp["カルテ枚数_前年"])

    # 表示列（店舗別の明細）
    show = [
        "店舗名","月",
        "総売上_前年","総売上_今年","総売上増減率%",
        "総来院数_前年","総来院数_今年","総来院数増減率%",
        "保険新患_前年","保険新患_今年","保険新患増減率%",
        "自由新患_前年","自由新患_今年","自由新患増減率%",
        "LTV_前年","LTV_今年","LTV増減率%",
        "カルテ枚数_前年","カルテ枚数_今年","カルテ枚数増減率%",
    ]
    detail_cols = [c for c in show if c in comp.columns]
    # ★追加：上のサマリーの列順を保存（下の「集計対象…一覧」で再利用）
    st.session_state["cols_allstore_summary"] = detail_cols

    comp_disp = comp[detail_cols].sort_values(["店舗名","月"])

    # 合計計算に使う列（増減率は除外）
    sum_targets = [c for c in detail_cols if c not in ("店舗名","月") and not str(c).endswith("増減率%")]

    # --- 店舗明細（増減率以外は整数カンマ）
    comp_disp = fmt_comma_int(comp_disp, sum_targets)

    # --- 月ごとの合計（月合計）
    agg_base = comp[["月"] + sum_targets].copy()
    mon_sum = agg_base.groupby("月", as_index=False).sum(numeric_only=True)
    for base in ["総売上","総来院数","保険新患","自由新患","LTV","カルテ枚数"]:
        a, b, r = f"{base}_今年", f"{base}_前年", f"{base}増減率%"
        if a in mon_sum.columns and b in mon_sum.columns:
            mon_sum[r] = _rate(mon_sum[a], mon_sum[b])
    mon_disp = mon_sum.copy()
    mon_disp.insert(0, "店舗名", mon_disp["月"].astype(int).astype(str) + "月合計")
    mon_disp = fmt_comma_int(mon_disp, [c for c in mon_disp.columns if c in sum_targets])
    mon_disp = mon_disp[detail_cols]

    # --- 総合計（= 月合計の合算）
    total_raw = mon_sum[sum_targets].sum(numeric_only=True)
    total_rec = {"店舗名":"総合計","月":"総計"}
    for k, v in total_raw.items():
        total_rec[k] = v
    for base in ["総売上","総来院数","保険新患","自由新患","LTV","カルテ枚数"]:
        a, b, r = f"{base}_今年", f"{base}_前年", f"{base}増減率%"
        if a in total_rec and b in total_rec:
            total_rec[r] = _rate(pd.Series([total_rec[a]]), pd.Series([total_rec[b]])).iloc[0]
    total_disp = pd.DataFrame([total_rec])
    total_disp = fmt_comma_int(total_disp, [c for c in sum_targets if c in total_disp.columns])
    total_disp = total_disp[detail_cols]

    final_disp = pd.concat([comp_disp, mon_disp, total_disp], ignore_index=True)
    st.dataframe(final_disp, use_container_width=True)
else:
    st.info("先に『売上管理表』を読み込んでください。")

# 3) 予実（売上・経費：目標/実績）
st.header("📈 予実（売上・経費：目標/実績）")
files_yj = st.file_uploader("📂 予実管理ファイル（グループ単位・複数可）…注意！読み取りに10分くらいかかります", type=["xlsx"], accept_multiple_files=True, key="yj_files")
yj_df = load_yj(files_yj) if files_yj else pd.DataFrame()
# ③ 店名統一（予実）
if not yj_df.empty:
    yj_df = unify_store_names(yj_df)
if yj_df.empty:
    st.caption("予実管理のブック（各シート＝店舗）を入れると、年・月×店舗の目標/実績（売上・経費）と粗利が見られます。")
else:
    y_opts = sorted(yj_df["年"].dropna().unique().tolist())
    c1, c2 = st.columns(2)
    with c1:
        sel_yj_y = st.selectbox("年（予実）", y_opts, index=len(y_opts)-1)
    with c2:
        m_opts = sorted(yj_df.query("年 == @sel_yj_y")["月"].dropna().unique().tolist())
        sel_yj_m = st.selectbox("月（予実）", m_opts, index=len(m_opts)-1 if m_opts else 0)

    yj_m = yj_df.query("年 == @sel_yj_y & 月 == @sel_yj_m").copy()
    if not yj_m.empty:
        for c in ["粗利率目標", "粗利率実績"]:
            if c in yj_m.columns:
                yj_m[c] = pd.to_numeric(yj_m[c], errors="coerce").round(1)
        money_cols = ["売上目標","売上実績","経費目標","経費実績","粗利目標","粗利実績"]
        yj_m = fmt_comma_int(yj_m, money_cols)
        show_cols = ["店舗名","売上目標","売上実績","経費目標","経費実績","粗利目標","粗利率目標","粗利実績","粗利率実績"]
        present = [c for c in show_cols if c in yj_m.columns]
        st.dataframe(yj_m[present].sort_values("売上実績", ascending=False, key=lambda s: pd.to_numeric(s.str.replace(",",""), errors="coerce")),
                     use_container_width=True)

    st.markdown("#### 📅 月ごとの一覧（年合計つき）")
# ---- 月ごとの一覧（年合計つき）※合計行は合計値から粗利・粗利率を再計算 ----
    yj_y = yj_df.query("年 == @sel_yj_y").copy()

    agg_cols = ["売上目標", "売上実績", "経費目標", "経費実績"]
    mon = yj_y.groupby("月", as_index=False)[agg_cols].sum(numeric_only=True)

    # 月別の粗利・粗利率
    mon["粗利目標"]   = mon["売上目標"] - mon["経費目標"]
    mon["粗利実績"]   = mon["売上実績"] - mon["経費実績"]
    mon["粗利率目標"] = (mon["粗利目標"] / mon["売上目標"] * 100).where(mon["売上目標"] != 0).round(1)
    mon["粗利率実績"] = (mon["粗利実績"] / mon["売上実績"] * 100).where(mon["売上実績"] != 0).round(1)

    # --- 合計行（粗利は「合計売上 - 合計経費」で再計算、粗利率も合計から算出） ---
    tot_sales_goal   = mon["売上目標"].sum()
    tot_sales_actual = mon["売上実績"].sum()
    tot_exp_goal     = mon["経費目標"].sum()
    tot_exp_actual   = mon["経費実績"].sum()

    tot_row = {
        "月": "合計",
        "売上目標":   tot_sales_goal,
        "売上実績":   tot_sales_actual,
        "経費目標":   tot_exp_goal,
        "経費実績":   tot_exp_actual,
    }
    # 粗利（合計値から算出）
    tot_row["粗利目標"]   = tot_sales_goal   - tot_exp_goal
    tot_row["粗利実績"]   = tot_sales_actual - tot_exp_actual
    # 粗利率（合計値から算出）
    tot_row["粗利率目標"] = round((tot_row["粗利目標"] / tot_sales_goal   * 100), 1) if tot_sales_goal   else None
    tot_row["粗利率実績"] = round((tot_row["粗利実績"] / tot_sales_actual * 100), 1) if tot_sales_actual else None

    mon_total = pd.concat([mon.sort_values("月"), pd.DataFrame([tot_row])], ignore_index=True)

    # 表示整形
    money_cols = ["売上目標","売上実績","経費目標","経費実績","粗利目標","粗利実績"]
    mon_total = fmt_comma_int(mon_total, money_cols)
    # 粗利率列は小数1桁のまま（None は空表示にしたい場合は下記を使用）
    for c in ["粗利率目標","粗利率実績"]:
        if c in mon_total.columns:
            mon_total[c] = mon_total[c].apply(lambda x: "" if pd.isna(x) else f"{x:.1f}")

    st.dataframe(mon_total, use_container_width=True, hide_index=True)


# 4) 個人生産性（店舗別ファイル・複数対応）
st.header("👤 個人生産性（店舗別ファイル・複数対応）")
files_person = st.file_uploader("📂 個人生産性ファイル（各店舗・各月のシート）", type=["xlsx"], accept_multiple_files=True, key="person_files")
pp_df, pp_logs = load_person_productivity(files_person) if files_person else (pd.DataFrame(), [])
# ③ 店名統一（個人生産性）
if not pp_df.empty:
    pp_df = unify_store_names(pp_df)
if pp_logs:
    with st.expander("📝 個人生産性 読み取りログ", expanded=False):
        for line in pp_logs:
            st.write(line)

if pp_df.empty:
    st.caption("各ブックの月シートから『稼働数／目標値/日／目標値/月／進捗／達成率』を読み取ります。")
else:
    # --- 年/月候補を正規化して作成 ---
    yser = pd.to_numeric(pp_df["年"], errors="coerce")
    yser = yser[(yser >= 2000) & (yser <= 2100)].astype(int)
    mser = pd.to_numeric(pp_df["月"], errors="coerce")
    mser = mser[(mser >= 1) & (mser <= 12)].astype(int)

    year_opts  = sorted(yser.unique().tolist())
    month_opts = sorted(mser.unique().tolist())

    c1, c2 = st.columns(2)
    with c1:
        sel_year = st.selectbox(
            "年（個人生産性）",
            year_opts,
            index=len(year_opts)-1,
            key="pp_year_select_main"      # ← 固有キー
        )
    with c2:
        idx_m = len(month_opts) - 1 if month_opts else 0
        sel_month = st.selectbox(
            "月（個人生産性）",
            month_opts,
            index=idx_m,
            key="pp_month_select_main",     # ← 固有キー
            disabled=(len(month_opts) == 0)
        )

    # --- フィルタ＆表示は 1 回だけ ---
    view = pp_df.query("年 == @sel_year & 月 == @sel_month").copy()

    # === 個人一覧を「担当≠院」と「院のみ」に分けて表示（派生列フル版） ===
    def _ensure_pp_derived(df: pd.DataFrame) -> pd.DataFrame:
        """日割平均 / 必要稼働数 / 必要稼働数の差 / 不足金額 / 日割達成率 を防弾で補完"""
        import numpy as np
        num = lambda c: pd.to_numeric(df.get(c), errors="coerce")

        # 日割平均 = 進捗 ÷ 稼働数（整数）
        if "日割平均" not in df.columns or df["日割平均"].isna().all():
            df["日割平均"] = (num("進捗") / num("稼働数")).round(0)

        # 必要稼働数 = ceil(目標値_月 ÷ 日割平均)（0/NaNはNaN）
        if "必要稼働数" not in df.columns or df["必要稼働数"].isna().all():
            den = pd.to_numeric(df["日割平均"], errors="coerce")
            df["必要稼働数"] = np.ceil(num("目標値_月") / den)
            df.loc[~np.isfinite(df["必要稼働数"]), "必要稼働数"] = np.nan

        # 必要稼働数の差 = 必要稼働数 − 稼働数（不足分のみ。マイナス→0）
        df["必要稼働数の差"] = (
            pd.to_numeric(df["必要稼働数"], errors="coerce") - num("稼働数")
        ).clip(lower=0)

        # 不足金額 = 目標値_月 − 進捗（不足分のみ。マイナス→0）
        df["不足金額"] = (num("目標値_月") - num("進捗")).clip(lower=0)

        # 日割達成率 = 日割平均 ÷ 目標値_日 ×100（無ければ補完）
        if "日割達成率" not in df.columns or df["日割達成率"].isna().all():
            df["日割達成率"] = (pd.to_numeric(df["日割平均"], errors="coerce") / num("目標値_日") * 100)

        # 月の達成率は常に再計算（進捗 ÷ 目標値_月 ×100）
        df["達成率"] = (num("進捗") / num("目標値_月") * 100)

        return df

    # 表示列（個人・院 共通）
    show_cols = [
        "年","月","店舗名","担当",
        "稼働数","必要稼働数","必要稼働数の差",
        "目標値_日","日割平均","日割達成率",
        "目標値_月","進捗","不足金額","達成率"
    ]
    money_cols = ["稼働数","必要稼働数","必要稼働数の差","目標値_日","日割平均","目標値_月","進捗","不足金額"]
    # ★追加：上の個人生産性の列順を保存（下の「店舗×月 指定」一覧で再利用）
    st.session_state["cols_person_view"] = show_cols
    # -------- 1) 担当 ≠ 「院」 --------
    staff = view.query("担当 != '院'").copy()
    if not staff.empty:
        staff = _ensure_pp_derived(staff)
        staff = fmt_comma_int(staff, money_cols)
        staff = fmt_percent(staff, ["達成率", "日割達成率"])
        st.dataframe(
            staff[[c for c in show_cols if c in staff.columns]].sort_values(["店舗名","担当"]),
            use_container_width=True, hide_index=True
        )
    else:
        st.info("担当（院以外）の行はありません。")

    # -------- 2) 「院」だけ --------
    inn = view.query("担当 == '院'").copy()
    if not inn.empty:
        st.markdown("##### 院（合計）")
        inn = _ensure_pp_derived(inn)
        inn = fmt_comma_int(inn, money_cols)
        inn = fmt_percent(inn, ["達成率", "日割達成率"])
        st.dataframe(
            inn[[c for c in show_cols if c in inn.columns]].sort_values(["店舗名"]),
            use_container_width=True, hide_index=True
        )

    # === 月ごとのサマリ（院のみ：目標値_月・進捗・達成率） ==========================
    st.markdown("#### 🗓 月ごとのサマリ（院のみ）")
    base = pp_df.query("年 == @sel_year & 担当 == '院'").copy()
    if base.empty:
        st.info("この年の『院』データがありません。")
    else:
        m = pd.to_numeric(base["目標値_月"], errors="coerce")
        p = pd.to_numeric(base["進捗"],      errors="coerce")
        g = base.assign(_m=m, _p=p)\
                .groupby("月", as_index=False)\
                .agg(目標値_月=("_m","sum"), 進捗=("_p","sum"))
        g["達成率"] = (g["進捗"] / g["目標値_月"] * 100)

        # 合計行（合計から再計算）
        tot_m, tot_p = g["目標値_月"].sum(), g["進捗"].sum()
        tot = {"月":"合計", "目標値_月":tot_m, "進捗":tot_p,
               "達成率": (tot_p / tot_m * 100) if tot_m else None}
        out = pd.concat([g, pd.DataFrame([tot])], ignore_index=True)

        out = fmt_comma_int(out, ["目標値_月","進捗"])
        out = fmt_percent(out, ["達成率"])
        st.dataframe(out[["月","目標値_月","進捗","達成率"]],
                     use_container_width=True, hide_index=True)

   

# 5) 店舗選択（シンプル）
# === 🏪 店舗情報（KPIカード／稼働×効率／来院・顧客価値） =========================
st.header("🏪 店舗情報")

# --- 店舗の候補（どのDFにも出てくる店舗をまとめる）
stores_all = []
for _df in [sales_df, yj_df, pp_df]:
    if not _df.empty and "店舗名" in _df.columns:
        stores_all.extend(_df["店舗名"].dropna().astype(str).tolist())
stores_all = sorted(pd.unique(pd.Series(stores_all)))
if not stores_all:
    st.info("店舗情報を表示するには、先に売上管理・予実・個人生産性のいずれかを読み込んでください。")
else:
    c0, c1, c2 = st.columns([2,1,1])
    with c0:
        sel_store = st.selectbox("店舗を選択（店舗情報）", stores_all, key="store_info_select")


    # --- 年月の候補：予実優先→売上管理→個人生産性から推定
    def _ym_opts(df):
        # None / 空DF / 必須カラム欠落はスキップ
        if df is None or df.empty or not {"店舗名", "年", "月"}.issubset(df.columns):
            return []
        sub = df.loc[df["店舗名"] == sel_store, ["年", "月"]].copy()
        # 数値化してから欠損を除外
        sub["年"] = pd.to_numeric(sub["年"], errors="coerce")
        sub["月"] = pd.to_numeric(sub["月"], errors="coerce")
        sub = sub.dropna()
        # (年, 月) のタプルにして返却（重複除去）
        return list(set(zip(sub["年"].astype(int), sub["月"].astype(int))))

    # ★ ここを “or 連鎖” ではなく合集合に変更
    ym_set = set(_ym_opts(yj_df)) | set(_ym_opts(sales_df)) | set(_ym_opts(pp_df))
    ym = sorted(ym_set)  # [(年, 月), ...] を年→月で昇順

    c1, c2 = st.columns(2)

    if ym:
        # 実データから候補を作成（前回選択を優先）
        y_opts = sorted({y for y, _ in ym})
        prev_y = st.session_state.get("store_info_year")
        y_idx = y_opts.index(prev_y) if prev_y in y_opts else len(y_opts) - 1
        with c1:
            sel_year = st.selectbox("年", y_opts, index=y_idx, key="store_info_year")

        m_opts = sorted({m for y, m in ym if y == sel_year})
        prev_m = st.session_state.get("store_info_month")
        m_idx = m_opts.index(prev_m) if prev_m in m_opts else len(m_opts) - 1
        with c2:
            sel_month = st.selectbox("月", m_opts, index=m_idx, key="store_info_month")
    else:
        # ★未読込などで年月が見つからない時：安全なデフォルトで“空表示”
        y_fallback, m_fallback = safe_year_month_for_sales(
            st.session_state.get("store_info_year"),
            st.session_state.get("store_info_month"),
            sales_df if 'sales_df' in locals() else None
        )
        st.warning("この店舗の年月データが見つかりません（未読込の場合は空表示）。")
        with c1:
            sel_year  = st.selectbox("年", [y_fallback], index=0, key="store_info_year", disabled=True)
        with c2:
            sel_month = st.selectbox("月", [m_fallback], index=0, key="store_info_month", disabled=True)


    
# ===================== セクション1：患者分析（売上管理表 由来） =====================
    # ─────────────────────────────────────────────
    # LTV / カルテ枚数 / 総売上 / 総来院数（前年同月比カード）
    # ─────────────────────────────────────────────

    # 1) その店舗の「最新年（latest）×選択月（sel_month）」と「前年同月」を揃える
    # ここは選択値をそのまま使う。年月候補が無い時だけフォールバック。
    if ym:
        yr, mo = int(sel_year), int(sel_month)
    else:
        yr, mo = safe_year_month_for_sales(
            st.session_state.get("store_info_year"),
            st.session_state.get("store_info_month"),
            sales_df if 'sales_df' in locals() else None
        )
    
    prev = yr - 1

    def _pick(df, value_col):
        """（店舗・年・月）で 1 値を引くヘルパー。無い場合 None。"""
        try:
            v = df.query("店舗名 == @sel_store & 年 == @yr & 月 == @mo")[value_col].iloc[0]
            return float(v) if pd.notna(v) else None
        except Exception:
            return None

    def _pick_prev(df, value_col):
        try:
            v = df.query("店舗名 == @sel_store & 年 == @prev & 月 == @mo")[value_col].iloc[0]
            return float(v) if pd.notna(v) else None
        except Exception:
            return None

    # 2) 必要な4指標の「今年/前年」を取得
    #   LTV / カルテ枚数 は元の集約DF（ltv_df, card_df）から
    cur_ltv  = _pick(ltv_df,  "LTV")   if not ltv_df.empty  else None
    prv_ltv  = _pick_prev(ltv_df, "LTV") if not ltv_df.empty else None

    cur_card = _pick(card_df, "カルテ枚数")   if not card_df.empty else None
    prv_card = _pick_prev(card_df, "カルテ枚数") if not card_df.empty else None

    #   総売上 / 総来院数 は sales_df を店舗×年×月で合算して利用
    agg_cols = [c for c in ["総売上", "総来院数", "保険新患", "自由新患"] if c in sales_df.columns]
    _sales_mon = (
        sales_df.groupby(["店舗名","年","月"], as_index=False)[agg_cols].sum(numeric_only=True)
    ) if (not sales_df.empty and agg_cols) else pd.DataFrame(columns=["店舗名","年","月"] + agg_cols)

    def _pick_sales(col):
        try:
            v = _sales_mon.query("店舗名 == @sel_store & 年 == @yr & 月 == @mo")[col].iloc[0]
            return float(v) if pd.notna(v) else None
        except Exception:
            return None

    def _pick_sales_prev(col):
        try:
            v = _sales_mon.query("店舗名 == @sel_store & 年 == @prev & 月 == @mo")[col].iloc[0]
            return float(v) if pd.notna(v) else None
        except Exception:
            return None

    cur_sales = _pick_sales("総売上")
    prv_sales = _pick_sales_prev("総売上")

    cur_vis   = _pick_sales("総来院数")
    prv_vis   = _pick_sales_prev("総来院数")

    # 3) カード描画
    st.markdown("##### LTV / カルテ枚数 / 総売上 / 総来院数（前年同月比）")
    c1, c2, c3, c4 = st.columns(4)
    _metric_card(c1, "LTV",        cur_ltv,  prv_ltv,  "円")
    _metric_card(c2, "カルテ枚数",  cur_card, prv_card, "枚")
    _metric_card(c3, "総売上",      cur_sales, prv_sales, "円")
    _metric_card(c4, "総来院数",    cur_vis,   prv_vis,   "人")
    def _pick_sales(col):
        try:
            v = _sales_mon.query("店舗名 == @sel_store & 年 == @yr & 月 == @mo")[col].iloc[0]
            return float(v) if pd.notna(v) else None
        except Exception:
            return None

    def _pick_sales_prev(col):
        try:
            v = _sales_mon.query("店舗名 == @sel_store & 年 == @prev & 月 == @mo")[col].iloc[0]
            return float(v) if pd.notna(v) else None
        except Exception:
            return None

    cur_ins = _pick_sales("保険新患") if "保険新患" in _sales_mon.columns else None
    prv_ins = _pick_sales_prev("保険新患") if "保険新患" in _sales_mon.columns else None
    cur_sp  = _pick_sales("自由新患") if "自由新患" in _sales_mon.columns else None
    prv_sp  = _pick_sales_prev("自由新患") if "自由新患" in _sales_mon.columns else None

    c5, c6 = st.columns(2)
    _metric_card(c5, "保険新患", cur_ins, prv_ins, "人")
    _metric_card(c6, "自由新患", cur_sp,  prv_sp,  "人")

    # ========== 月ごとの前年比較（総売上・総来院数・保険新患・自由新患：横並び） ==========
    st.markdown("### 📈 月ごとの前年比較")

    if (_sales_mon is None) or _sales_mon.empty or (sel_store is None):
        st.info("売上管理の月次集計が不足しています。ファイルを読み込んでください。")
    else:
        latest_year = int(_sales_mon[_sales_mon["店舗名"] == sel_store]["年"].max())
        prev_year   = latest_year - 1

        metric_defs = [
            ("総売上",   "総売上",   "円"),
            ("総来院数", "総来院数", "人"),
            ("保険新患", "保険新患", "人"),
            ("自由新患", "自由新患", "人"),
        ]
        metrics = [(k, t, u) for (k, t, u) in metric_defs if k in _sales_mon.columns]

        if not metrics:
            st.info("比較可能な指標がありません（総売上 / 総来院数 / 保険新患 / 自由新患 の列が不足）。")
        else:
            import altair as alt

            def yoy_grouped(metric_col: str, title: str, unit: str):
                base = _sales_mon.query(
                    "店舗名 == @sel_store and 年 in (@latest_year, @prev_year)"
                )[["年", "月", metric_col]].copy()

                base[metric_col] = pd.to_numeric(base[metric_col], errors="coerce").fillna(0)
                base["年"] = base["年"].astype(str)  # 色・オフセット用に文字列化
                month_order = list(range(1, 12 + 1))

                chart = (
                    alt.Chart(base)
                    .mark_bar()
                    .encode(
                        x=alt.X("月:O", title="月", sort=month_order, axis=alt.Axis(labelAngle=0)),
                        # ← 年ごとの棒を横にずらす（グループ化のキモ）
                        xOffset=alt.XOffset("年:N"),
                        color=alt.Color("年:N", title="年度"),
                        y=alt.Y(f"{metric_col}:Q", title=f"{title}（{unit}）"),
                        tooltip=[
                            alt.Tooltip("年:N", title="年度"),
                            alt.Tooltip("月:O", title="月"),
                            alt.Tooltip(f"{metric_col}:Q", title=title, format=",.0f"),
                        ],
                    )
                    .properties(height=260)
                )

                st.altair_chart(chart, use_container_width=True)

            rows = [metrics[i:i+2] for i in range(0, len(metrics), 2)]
            for row in rows:
                cols = st.columns(len(row))
                for (col, (mkey, mtitle, unit)) in zip(cols, row):
                    with col:
                        yoy_grouped(mkey, mtitle, unit)
    # ========== 集計対象 全店舗サマリー（今年 vs 前年：月×店舗） ==========
    st.markdown("### 🧾 集計対象になった 全店舗サマリー（今年 vs 前年）")

    if ('_sales_mon' not in locals()) or _sales_mon is None or _sales_mon.empty:
        st.info("売上管理の月次集計が不足しています。ファイルを読み込んでください。")
    else:
        # 直近の年を軸に前年比較
        latest_year_all = int(_sales_mon["年"].max())
        prev_year_all   = latest_year_all - 1

        # 使う基本指標（存在するものだけ）
        base_cols = [c for c in ["総売上","総来院数","保険新患","自由新患"] if c in _sales_mon.columns]

        cur = _sales_mon.query("年 == @latest_year_all")[["店舗名","月"] + base_cols].copy()
        prv = _sales_mon.query("年 == @prev_year_all")[["店舗名","月"] + base_cols].copy()
        cur = cur.add_suffix("_今年").rename(columns={"店舗名_今年":"店舗名","月_今年":"月"})
        prv = prv.add_suffix("_前年").rename(columns={"店舗名_前年":"店舗名","月_前年":"月"})

        comp = pd.merge(cur, prv, on=["店舗名","月"], how="outer")

        # LTV / カルテ枚数（上段と同じ要領で結合：あれば付与）
        if 'ltv_df' in locals() and ltv_df is not None and not ltv_df.empty:
            ltv_c = (ltv_df[ltv_df["年"]==latest_year_all][["店舗名","月","LTV"]]
                    .rename(columns={"LTV":"LTV_今年"}))
            ltv_p = (ltv_df[ltv_df["年"]==prev_year_all][["店舗名","月","LTV"]]
                    .rename(columns={"LTV":"LTV_前年"}))
            comp = comp.merge(ltv_c, on=["店舗名","月"], how="left").merge(ltv_p, on=["店舗名","月"], how="left")

        if 'card_df' in locals() and card_df is not None and not card_df.empty:
            cd_c = (card_df[card_df["年"]==latest_year_all][["店舗名","月","カルテ枚数"]]
                    .rename(columns={"カルテ枚数":"カルテ枚数_今年"}))
            cd_p = (card_df[card_df["年"]==prev_year_all][["店舗名","月","カルテ枚数"]]
                    .rename(columns={"カルテ枚数":"カルテ枚数_前年"}))
            comp = comp.merge(cd_c, on=["店舗名","月"], how="left").merge(cd_p, on=["店舗名","月"], how="left")

        # 増減率（上段と同じ関数でOK）
        def _rate(a, b):
            a = pd.to_numeric(a, errors="coerce")
            b = pd.to_numeric(b, errors="coerce")
            den = b.replace(0, pd.NA)
            return ((a - b) / den * 100).astype("Float64").round(1)

        for base in ["総売上","総来院数","保険新患","自由新患","LTV","カルテ枚数"]:
            a, p = f"{base}_今年", f"{base}_前年"
            if (a in comp.columns) and (p in comp.columns):
                comp[f"{base}増減率%"] = _rate(comp[a], comp[p])

        # ---- 選択店舗だけに絞る（店舗情報セレクタと連動） ----
        sel_store = st.session_state.get("store_info_select", st.session_state.get("info_store"))
        if sel_store:
            comp = comp[comp["店舗名"] == sel_store]

        # 上段サマリーの列順に揃える（存在する列のみ）
        cols_all = st.session_state.get("cols_allstore_summary")
        if cols_all:
            comp = comp.reindex(columns=[c for c in cols_all if c in comp.columns])

        # 金額・件数はカンマ整形（“_今年/前年”で終わる列）
        money_like = [c for c in comp.columns
                    if (c.endswith("_今年") or c.endswith("_前年")) and (not c.endswith("増減率%"))]
        comp_disp = fmt_comma_int(comp.copy(), money_like)

        st.dataframe(comp_disp.sort_values(["店舗名","月"]),
                    use_container_width=True, hide_index=True)


    st.markdown("### 👥 患者分析")

# この店舗で患者分析DFに存在する最新年を決定
    years_in_pa = []
    for _df in (gender_df, reason_df, age_df):
        if not _df.empty:
            years_in_pa.extend(
                _df.query("店舗名 == @sel_store")["年"].dropna().astype(int).tolist()
            )
    if years_in_pa:
        latest = max(years_in_pa)
        prev   = latest - 1
    else:
        latest = int(sel_year); prev = latest - 1  # フォールバック

    # 最新年の構成（横棒＋割合%）— app - コピー.py と同じ見た目
    c1, c2, c3 = st.columns(3)
    with c1: plot_pivot(gender_df, sel_store, latest, "男女比率")
    with c2: plot_pivot(reason_df, sel_store, latest, "来院動機")
    with c3: plot_pivot(age_df,    sel_store, latest, "年齢比率")

    # 前年 vs 今年（棒の比較＋増減テーブル）
    st.markdown("#### 📊 前年比較")
    plot_cat_yoy(reason_df, sel_store, latest, prev, "来院動機")
    plot_cat_yoy(gender_df, sel_store, latest, prev, "男女比率")
    plot_cat_yoy(age_df,    sel_store, latest, prev, "年齢比率")
# =============================
# ① 予実（カード表示：店舗×年月）
# =============================
st.subheader("① 予実（店舗×月 指定：カード）")

if 'yj_df' not in locals():
    st.info("予実データが未読込です。上部の『予実管理ファイル』を読み込んでください。")
else:
    if yj_df.empty:
        st.info("予実データが空です。")
    else:
        # 店舗情報のセレクタと連動
        kpi_store = st.session_state.get("store_info_select")
        kpi_year  = st.session_state.get("store_info_year")
        kpi_month = st.session_state.get("store_info_month")

        # 念のためのフォールバック（初回読み込み時など）
        if kpi_store is None:
            kpi_store = sorted(yj_df["店舗名"].dropna().unique().tolist())[0]
        if kpi_year is None:
            kpi_year = int(yj_df["年"].dropna().max())
        if kpi_month is None:
            kpi_month = int(yj_df.query("年 == @kpi_year")["月"].dropna().max())

        yj_one = yj_df.query("店舗名 == @kpi_store & 年 == @kpi_year & 月 == @kpi_month").copy()
        if yj_one.empty:
            st.warning("該当する予実レコードがありません。")
        else:
            # 複数行あっても合算（安全）
            s_goal  = pd.to_numeric(yj_one.get("売上目標"), errors="coerce").sum()
            s_act   = pd.to_numeric(yj_one.get("売上実績"), errors="coerce").sum()
            e_goal  = pd.to_numeric(yj_one.get("経費目標"), errors="coerce").sum()
            e_act   = pd.to_numeric(yj_one.get("経費実績"), errors="coerce").sum()
            g_goal  = s_goal - e_goal
            g_act   = s_act - e_act

            def _rate(act, goal):
                if pd.isna(goal) or goal == 0: return None
                return float(act) / float(goal) * 100.0

            r_sales = _rate(s_act, s_goal)
            r_exp   = _rate(e_act, e_goal)
            r_gross = _rate(g_act, g_goal)

            def _fmt_int(v): 
                return "" if pd.isna(v) else f"{int(round(v)):,}"
            def _fmt_pct(v):
                return "" if (v is None or pd.isna(v)) else f"{v:.1f}%"

            col1, col2, col3 = st.columns(3)

            with col1:
                st.markdown("##### 売上")
                st.metric(
                    label=f"目標 {_fmt_int(s_goal)} 円",
                    value=f"{_fmt_int(s_act)} 円",
                    delta=f"達成率 {_fmt_pct(r_sales)}",
                    delta_color="normal" if (r_sales or 0) >= 100 else "inverse"
                )

            with col2:
                st.markdown("##### 経費")
                st.metric(
                    label=f"目標 {_fmt_int(e_goal)} 円",
                    value=f"{_fmt_int(e_act)} 円",
                    delta=f"達成率 {_fmt_pct(r_exp)}",
                    # 経費は“低いほど良い”だが、ここは達成率の表示ルールに合わせて通常色
                    delta_color="normal" if (r_exp or 0) <= 100 else "inverse"
                )

            with col3:
                st.markdown("##### 粗利")
                st.metric(
                    label=f"目標 {_fmt_int(g_goal)} 円",
                    value=f"{_fmt_int(g_act)} 円",
                    delta=f"達成率 {_fmt_pct(r_gross)}",
                    delta_color="normal" if (r_gross or 0) >= 100 else "inverse"
                )


    # ===================================
    # ② 個人生産（一覧：店舗×年月 の担当者別）
    # ===================================
    st.subheader("② 個人生産（店舗×月 指定：担当者別一覧）")

    if 'pp_df' not in locals():
        st.info("個人生産データが未読込です。上部の『個人生産性ファイル』を読み込んでください。")
    else:
        if pp_df.empty:
            st.info("個人生産データが空です。")
        else:
            # 店舗情報のセレクタと連動
            kpi_store_pp = st.session_state.get("store_info_select")
            kpi_year_pp  = st.session_state.get("store_info_year")
            kpi_month_pp = st.session_state.get("store_info_month")

            # 念のためのフォールバック
            if kpi_store_pp is None:
                kpi_store_pp = sorted(pp_df["店舗名"].dropna().unique().tolist())[0]
            if kpi_year_pp is None:
                kpi_year_pp = int(pp_df["年"].dropna().max())
            if kpi_month_pp is None:
                kpi_month_pp = int(pp_df.query("年 == @kpi_year_pp")["月"].dropna().max())

            view = pp_df.query(
                "店舗名 == @kpi_store_pp & 年 == @kpi_year_pp & 月 == @kpi_month_pp"
            ).copy()

            if view.empty:
                st.warning("該当する個人生産レコードがありません。")
            else:
                # 互換対応：昔の列名が来たら揃える
                if "不足稼働数" in view.columns and "必要稼働数の差" not in view.columns:
                    view = view.rename(columns={"不足稼働数": "必要稼働数の差"})

                # 上の個人生産性セクションで定義した補完関数を再利用
                view = _ensure_pp_derived(view)

                # 表示整形（数値 → カンマ、％ → xx.x）
                money_cols = [
                    "稼働数","必要稼働数","必要稼働数の差",
                    "目標値_日","日割平均","目標値_月","進捗","不足金額"
                ]
                view = fmt_comma_int(view, [c for c in money_cols if c in view.columns])
                view = fmt_percent(view, ["達成率", "日割達成率"])

                # 列順は“上の個人生産性”と同じにする（保存済み）
                cols_pp = st.session_state.get("cols_person_view")
                if cols_pp:
                    show_cols = [c for c in cols_pp if c in view.columns]
                else:
                    # フォールバック（基本同じ並び）
                    show_cols = [
                        "年","月","店舗名","担当",
                        "稼働数","必要稼働数","必要稼働数の差",
                        "目標値_日","日割平均","日割達成率",
                        "目標値_月","進捗","不足金額","達成率"
                    ]
                    show_cols = [c for c in show_cols if c in view.columns]

                st.dataframe(
                    view[show_cols].sort_values(["店舗名","担当"]),
                    use_container_width=True, hide_index=True
                )
